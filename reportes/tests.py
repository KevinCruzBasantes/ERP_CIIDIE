import io
import tempfile
from datetime import timedelta

import openpyxl
from django.test import TestCase, override_settings
from django.urls import reverse
from django.utils import timezone

from testing_comun import (
    crear_estudiante,
    crear_maquina,
    crear_operador,
    crear_tecnico,
)
from reportes.models import ReporteGenerado

# Los archivos generados van a un MEDIA_ROOT temporal para no ensuciar media/
TEMP_MEDIA = tempfile.mkdtemp(prefix='test_media_reportes_')

XLSX_MIME = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

ENDPOINTS = [
    ('reporte_resumen',       'RESUMEN'),
    ('reporte_mantenimiento', 'MANTENIMIENTO'),
    ('reporte_produccion',    'PRODUCCION'),
    ('reporte_inventario',    'INVENTARIO'),
    ('reporte_seguridad',     'SEGURIDAD'),
    ('reporte_backup',        'BACKUP'),
]


@override_settings(MEDIA_ROOT=TEMP_MEDIA)
class GeneracionReportesTest(TestCase):

    def setUp(self):
        self.tecnico = crear_tecnico()
        self.maquina = crear_maquina()
        self.client.force_login(self.tecnico)

    def test_los_seis_reportes_generan_excel_valido(self):
        for url, tipo in ENDPOINTS:
            with self.subTest(reporte=tipo):
                respuesta = self.client.post(reverse(url))
                self.assertEqual(respuesta.status_code, 200)
                self.assertEqual(respuesta['Content-Type'], XLSX_MIME)
                # el workbook descargado abre sin errores
                wb = openpyxl.load_workbook(io.BytesIO(respuesta.content))
                self.assertGreaterEqual(len(wb.sheetnames), 1)
                # queda en el historial con el mismo archivo re-descargable
                reporte = ReporteGenerado.objects.filter(tipo=tipo).latest('fecha_generacion')
                self.assertEqual(reporte.generado_por, self.tecnico)
                with reporte.archivo.open('rb') as f:
                    self.assertEqual(f.read(), respuesta.content)

    def test_get_no_genera_solo_redirige(self):
        respuesta = self.client.get(reverse('reporte_resumen'))
        self.assertRedirects(respuesta, reverse('lista_reportes'))
        self.assertEqual(ReporteGenerado.objects.count(), 0)

    def test_el_periodo_del_filtro_queda_en_el_historial(self):
        hoy = timezone.now().date()
        inicio, fin = hoy - timedelta(days=30), hoy
        self.client.post(reverse('reporte_mantenimiento'), {
            'fecha_inicio': inicio.isoformat(), 'fecha_fin': fin.isoformat()})
        reporte = ReporteGenerado.objects.get(tipo='MANTENIMIENTO')
        self.assertEqual(reporte.fecha_inicio_periodo, inicio)
        self.assertEqual(reporte.fecha_fin_periodo, fin)

    def test_la_lista_muestra_el_historial(self):
        self.client.post(reverse('reporte_resumen'))
        respuesta = self.client.get(reverse('lista_reportes'))
        self.assertEqual(respuesta.status_code, 200)
        self.assertEqual(len(respuesta.context['reportes']), 1)
        self.assertEqual(len(respuesta.context['tarjetas']), 6)

    def test_operador_no_accede_a_reportes(self):
        self.client.force_login(crear_operador())
        self.assertEqual(self.client.get(reverse('lista_reportes')).status_code, 302)
        self.assertEqual(self.client.post(reverse('reporte_backup')).status_code, 302)
        self.assertEqual(ReporteGenerado.objects.count(), 0)

    def test_estudiante_no_accede_a_reportes(self):
        self.client.force_login(crear_estudiante())
        respuesta = self.client.get(reverse('lista_reportes'))
        self.assertRedirects(respuesta, reverse('dashboard_general'))


@override_settings(MEDIA_ROOT=TEMP_MEDIA)
class RespaldoCompletoTest(TestCase):

    def setUp(self):
        self.tecnico = crear_tecnico(username='resp')
        self.maquina = crear_maquina(nombre='Fresadora CNC')
        self.client.force_login(self.tecnico)

    def test_incluye_todas_las_tablas_del_sistema(self):
        respuesta = self.client.post(reverse('reporte_backup'))
        wb = openpyxl.load_workbook(io.BytesIO(respuesta.content))
        # una hoja por modelo de las 7 apps + índice
        for hoja in ('Índice', 'Usuario', 'Rol', 'Maquina', 'Pieza', 'Material',
                     'OrdenMantenimiento', 'Reserva', 'OrdenTrabajo', 'Alerta',
                     'CertificacionUsuario', 'ReporteGenerado'):
            self.assertIn(hoja, wb.sheetnames)
        self.assertGreaterEqual(len(wb.sheetnames), 25)

    def test_los_datos_reales_aparecen_en_el_respaldo(self):
        respuesta = self.client.post(reverse('reporte_backup'))
        wb = openpyxl.load_workbook(io.BytesIO(respuesta.content))
        nombres = [fila[0].value for fila in wb['Maquina'].iter_rows(min_row=2)]
        # la columna exacta depende del orden de campos; buscar en toda la hoja
        valores = {c.value for fila in wb['Maquina'].iter_rows(min_row=2) for c in fila}
        self.assertIn('Fresadora CNC', valores)

    def test_nunca_exporta_hashes_de_contrasenas(self):
        respuesta = self.client.post(reverse('reporte_backup'))
        wb = openpyxl.load_workbook(io.BytesIO(respuesta.content))
        encabezados = [c.value for c in wb['Usuario'][1]]
        self.assertNotIn('password', encabezados)
        contenido = respuesta.content
        self.assertNotIn(self.tecnico.password.encode(), contenido)
