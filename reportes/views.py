from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils import timezone
from django.contrib import messages
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from .models import ReporteGenerado
from inventario.models import Material
from mantenimiento.models import Mantenimiento
from tpm.models import InspeccionDiaria, RegistroOEE
from django.utils import timezone as tz


def estilo_encabezado(ws, fila, columnas):
    """Aplica estilo de encabezado a una fila."""
    fill = PatternFill(start_color='1E2333', end_color='1E2333', fill_type='solid')
    font = Font(bold=True, color='E8A020', size=10)
    for col, texto in enumerate(columnas, 1):
        cell = ws.cell(row=fila, column=col, value=texto)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')


@login_required(login_url='login')
def lista_reportes(request):
    reportes = ReporteGenerado.objects.select_related('generado_por').all()
    context = {
        'reportes': reportes,
        'tipos': ReporteGenerado.TIPOS,
    }
    return render(request, 'reportes/lista_reportes.html', context)


@login_required(login_url='login')
def generar_inventario(request):
    if request.method != 'POST':
        return redirect('lista_reportes')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Inventario'

    encabezados = ['Código', 'Nombre', 'Tipo', 'Unidad', 'Stock actual',
                   'Stock mínimo', 'Estado stock', 'Proveedor', 'Costo unitario']
    estilo_encabezado(ws, 1, encabezados)

    materiales = Material.objects.filter(activo=True)
    for i, m in enumerate(materiales, 2):
        ws.cell(row=i, column=1, value=m.codigo)
        ws.cell(row=i, column=2, value=m.nombre)
        ws.cell(row=i, column=3, value=m.get_tipo_display())
        ws.cell(row=i, column=4, value=m.unidad_medida)
        ws.cell(row=i, column=5, value=float(m.stock_actual))
        ws.cell(row=i, column=6, value=float(m.stock_minimo))
        ws.cell(row=i, column=7, value='Stock bajo' if m.stock_bajo else 'OK')
        ws.cell(row=i, column=8, value=m.proveedor or '—')
        ws.cell(row=i, column=9, value=float(m.costo_unitario))

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    reporte = ReporteGenerado.objects.create(
        tipo='INVENTARIO',
        generado_por=request.user,
        fecha_inicio_periodo=timezone.now().date(),
        fecha_fin_periodo=timezone.now().date(),
    )

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="inventario_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


@login_required(login_url='login')
def generar_mantenimiento(request):
    if request.method != 'POST':
        return redirect('lista_reportes')

    fecha_inicio = request.POST.get('fecha_inicio')
    fecha_fin = request.POST.get('fecha_fin')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Mantenimiento'

    encabezados = ['#', 'Máquina', 'Código', 'Tipo', 'Estado', 'Prioridad',
                   'Fecha programada', 'Fecha inicio', 'Fecha fin',
                   'Horas trabajo', 'Costo', 'Responsable', 'Descripción']
    estilo_encabezado(ws, 1, encabezados)

    qs = Mantenimiento.objects.select_related('maquina', 'responsable')
    if fecha_inicio:
        qs = qs.filter(fecha_programada__gte=fecha_inicio)
    if fecha_fin:
        qs = qs.filter(fecha_programada__lte=fecha_fin)

    for i, m in enumerate(qs, 2):
        ws.cell(row=i, column=1, value=m.pk)
        ws.cell(row=i, column=2, value=m.maquina.nombre)
        ws.cell(row=i, column=3, value=m.maquina.codigo)
        ws.cell(row=i, column=4, value=m.get_tipo_display())
        ws.cell(row=i, column=5, value=m.get_estado_display())
        ws.cell(row=i, column=6, value=m.get_prioridad_display())
        ws.cell(row=i, column=7, value=m.fecha_programada.strftime('%d/%m/%Y'))
        ws.cell(row=i, column=8, value=tz.localtime(m.fecha_inicio).strftime('%d/%m/%Y %H:%M') if m.fecha_inicio else '—')
        ws.cell(row=i, column=9, value=tz.localtime(m.fecha_fin).strftime('%d/%m/%Y %H:%M') if m.fecha_fin else '—')
        ws.cell(row=i, column=10, value=float(m.horas_trabajo))
        ws.cell(row=i, column=11, value=float(m.costo))
        ws.cell(row=i, column=12, value=m.responsable.get_full_name() if m.responsable else '—')
        ws.cell(row=i, column=13, value=m.descripcion)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    ReporteGenerado.objects.create(
        tipo='MANTENIMIENTO',
        generado_por=request.user,
        fecha_inicio_periodo=fecha_inicio or timezone.now().date(),
        fecha_fin_periodo=fecha_fin or timezone.now().date(),
    )

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="mantenimiento_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


@login_required(login_url='login')
def generar_inspecciones(request):
    if request.method != 'POST':
        return redirect('lista_reportes')

    fecha_inicio = request.POST.get('fecha_inicio')
    fecha_fin = request.POST.get('fecha_fin')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Inspecciones'

    encabezados = ['Fecha', 'Máquina', 'Inspector', 'Limpieza', 'Temperatura', 'Guardas',
                   'Emergencia', 'Ruidos', 'Vibraciones', 'Ítems específicos (catálogo)',
                   'Resultado', 'Observaciones']
    estilo_encabezado(ws, 1, encabezados)

    qs = InspeccionDiaria.objects.select_related('maquina', 'inspector').prefetch_related('respuestas_checklist__item')
    if fecha_inicio:
        qs = qs.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        qs = qs.filter(fecha__lte=fecha_fin)

    si_no = lambda v: 'Sí' if v else 'No'
    ok_falla = lambda v: 'OK' if v else 'FALLA'

    for i, insp in enumerate(qs, 2):
        items_especificos = '; '.join(
            f"{r.item.nombre}: {'OK' if r.ok else 'FALLA'}" for r in insp.respuestas_checklist.all()
        ) or '—'
        ws.cell(row=i, column=1, value=insp.fecha.strftime('%d/%m/%Y'))
        ws.cell(row=i, column=2, value=insp.maquina.nombre)
        ws.cell(row=i, column=3, value=insp.inspector.get_full_name() if insp.inspector else '—')
        ws.cell(row=i, column=4, value=ok_falla(insp.limpieza_area_ok))
        ws.cell(row=i, column=5, value=ok_falla(insp.temperatura_normal))
        ws.cell(row=i, column=6, value=ok_falla(insp.guardas_seguridad_ok))
        ws.cell(row=i, column=7, value=ok_falla(insp.boton_emergencia_ok))
        ws.cell(row=i, column=8, value=si_no(insp.ruidos_anormales))
        ws.cell(row=i, column=9, value=si_no(insp.vibraciones_anormales))
        ws.cell(row=i, column=10, value=items_especificos)
        ws.cell(row=i, column=11, value='Aprobada' if insp.aprobada else 'FALLIDA')
        ws.cell(row=i, column=12, value=insp.observaciones or '—')

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 16

    ReporteGenerado.objects.create(
        tipo='TPM_INSPECCIONES',
        generado_por=request.user,
        fecha_inicio_periodo=fecha_inicio or timezone.now().date(),
        fecha_fin_periodo=fecha_fin or timezone.now().date(),
    )

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="inspecciones_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


@login_required(login_url='login')
def generar_oee(request):
    if request.method != 'POST':
        return redirect('lista_reportes')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'OEE'

    encabezados = ['Período', 'Máquina', 'Código', 'Disponibilidad %',
                   'Rendimiento %', 'Calidad %', 'OEE %', 'Clasificación']
    estilo_encabezado(ws, 1, encabezados)

    registros = RegistroOEE.objects.select_related('maquina').all()

    for i, r in enumerate(registros, 2):
        oee = float(r.oee)
        ws.cell(row=i, column=1, value=f"{r.mes:02d}/{r.anio}")
        ws.cell(row=i, column=2, value=r.maquina.nombre)
        ws.cell(row=i, column=3, value=r.maquina.codigo)
        ws.cell(row=i, column=4, value=float(r.disponibilidad))
        ws.cell(row=i, column=5, value=float(r.rendimiento))
        ws.cell(row=i, column=6, value=float(r.calidad))
        ws.cell(row=i, column=7, value=oee)
        ws.cell(row=i, column=8, value='World class' if oee >= 85 else 'Aceptable' if oee >= 60 else 'Mejorar')

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    ReporteGenerado.objects.create(
        tipo='TPM_OEE',
        generado_por=request.user,
        fecha_inicio_periodo=timezone.now().date(),
        fecha_fin_periodo=timezone.now().date(),
    )

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="oee_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response

@login_required(login_url='login')
def generar_reservas(request):
    if request.method != 'POST':
        return redirect('lista_reportes')

    fecha_inicio = request.POST.get('fecha_inicio')
    fecha_fin    = request.POST.get('fecha_fin')

    from reservas.models import Reserva
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Reservas'

    encabezados = ['#', 'Solicitante', 'Máquina', 'Código', 'Fecha',
                   'Hora inicio', 'Hora fin', 'Propósito', 'Estado',
                   'Autorizador', 'OT', 'T. planificado (min)', 'T. real (min)',
                   'T. paradas (min)', 'Unidades prod.', 'Unidades esp.']
    estilo_encabezado(ws, 1, encabezados)

    qs = Reserva.objects.select_related(
        'usuario', 'maquina', 'autorizador'
    ).prefetch_related('orden_trabajo')
    if fecha_inicio:
        qs = qs.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        qs = qs.filter(fecha__lte=fecha_fin)

    for i, r in enumerate(qs, 2):
        ot = getattr(r, 'orden_trabajo', None)
        ws.cell(row=i, column=1,  value=r.pk)
        ws.cell(row=i, column=2,  value=r.usuario.get_full_name() or r.usuario.username)
        ws.cell(row=i, column=3,  value=r.maquina.nombre)
        ws.cell(row=i, column=4,  value=r.maquina.codigo)
        ws.cell(row=i, column=5,  value=r.fecha.strftime('%d/%m/%Y'))
        ws.cell(row=i, column=6,  value=r.hora_inicio.strftime('%H:%M'))
        ws.cell(row=i, column=7,  value=r.hora_fin.strftime('%H:%M'))
        ws.cell(row=i, column=8,  value=r.get_proposito_display())
        ws.cell(row=i, column=9,  value=r.get_estado_display())
        ws.cell(row=i, column=10, value=r.autorizador.get_full_name() if r.autorizador else '—')
        ws.cell(row=i, column=11, value=f'OT-{ot.pk:04d}' if ot else '—')
        ws.cell(row=i, column=12, value=float(ot.tiempo_planificado_min) if ot else '—')
        ws.cell(row=i, column=13, value=float(ot.tiempo_real_min) if ot and ot.tiempo_real_min else '—')
        ws.cell(row=i, column=14, value=float(ot.tiempo_parada_min) if ot else '—')
        ws.cell(row=i, column=15, value=ot.unidades_producidas if ot else '—')
        ws.cell(row=i, column=16, value=ot.unidades_esperadas if ot else '—')

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    ReporteGenerado.objects.create(
        tipo='RESERVAS',
        generado_por=request.user,
        fecha_inicio_periodo=fecha_inicio or timezone.now().date(),
        fecha_fin_periodo=fecha_fin or timezone.now().date(),
    )

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reservas_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


@login_required(login_url='login')
def generar_pareto(request):
    if request.method != 'POST':
        return redirect('lista_reportes')

    fecha_inicio = request.POST.get('fecha_inicio')
    fecha_fin    = request.POST.get('fecha_fin')

    from reservas.models import RegistroParada
    from django.db.models import Count, Sum

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Pareto Paradas'

    encabezados = ['Código parada', 'Tipo', 'Categoría', 'Subsistema',
                   'Frecuencia', 'Duración total (min)', '% Frecuencia acumulada']
    estilo_encabezado(ws, 1, encabezados)

    qs = RegistroParada.objects.filter(activo=True).select_related('codigo_parada')
    if fecha_inicio:
        qs = qs.filter(orden_trabajo__reserva__fecha__gte=fecha_inicio)
    if fecha_fin:
        qs = qs.filter(orden_trabajo__reserva__fecha__lte=fecha_fin)

    # Agrupar por código de parada
    from collections import defaultdict
    agrupado = defaultdict(lambda: {'frecuencia': 0, 'duracion': 0, 'obj': None})
    for p in qs:
        key = p.codigo_parada.codigo if p.codigo_parada else 'SIN_CÓDIGO'
        agrupado[key]['frecuencia'] += 1
        agrupado[key]['duracion']   += float(p.duracion_minutos or 0)
        agrupado[key]['obj']         = p.codigo_parada

    # Ordenar por frecuencia descendente
    ordenado   = sorted(agrupado.items(), key=lambda x: x[1]['frecuencia'], reverse=True)
    total      = sum(v['frecuencia'] for _, v in ordenado)
    acumulado  = 0

    for i, (codigo, datos) in enumerate(ordenado, 2):
        acumulado += datos['frecuencia']
        pct        = round((acumulado / total * 100), 1) if total > 0 else 0
        obj        = datos['obj']
        ws.cell(row=i, column=1, value=codigo)
        ws.cell(row=i, column=2, value=obj.get_tipo_display() if obj else '—')
        ws.cell(row=i, column=3, value=obj.get_categoria_display() if obj else '—')
        ws.cell(row=i, column=4, value=obj.subsistema if obj else '—')
        ws.cell(row=i, column=5, value=datos['frecuencia'])
        ws.cell(row=i, column=6, value=round(datos['duracion'], 2))
        ws.cell(row=i, column=7, value=pct)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22

    ReporteGenerado.objects.create(
        tipo='TPM_PARETO',
        generado_por=request.user,
        fecha_inicio_periodo=fecha_inicio or timezone.now().date(),
        fecha_fin_periodo=fecha_fin or timezone.now().date(),
    )

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="pareto_paradas_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response