import datetime as dt
import io
from collections import defaultdict
from decimal import Decimal

from django.core.files.base import ContentFile
from django.db.models import Avg, F, Q
from django.http import HttpResponse
from django.utils import timezone
from django.utils import timezone as tz
import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font, PatternFill, Alignment

from .models import ReporteGenerado


# ─────────────────────────────────────────────────────────────────────────────
# Formateadores None-safe — para renderizar celdas sin reventar con FKs
# en NULL (SET_NULL) ni campos opcionales vacíos.
# ─────────────────────────────────────────────────────────────────────────────

def fmt_usuario(u):
    """Nombre completo (o username) de un Usuario, o '—' si el FK quedó en NULL."""
    if not u:
        return '—'
    return u.get_full_name() or u.username


def fmt_fecha(d):
    """DateField → 'dd/mm/aaaa' o '—'."""
    return d.strftime('%d/%m/%Y') if d else '—'


def fmt_fecha_hora(dt_):
    """DateTimeField → 'dd/mm/aaaa HH:MM' en hora local, o '—'."""
    return tz.localtime(dt_).strftime('%d/%m/%Y %H:%M') if dt_ else '—'


def si_no(v):
    return 'Sí' if v else 'No'


def fmt_num(x):
    """Decimal/int → float para la celda, o '—' si es None."""
    return float(x) if x is not None else '—'


def fmt_texto(t):
    """Texto opcional → el texto o '—' si está vacío."""
    return t if t else '—'


# ─────────────────────────────────────────────────────────────────────────────
# Construcción del workbook
# ─────────────────────────────────────────────────────────────────────────────

def estilo_encabezado(ws, fila, columnas):
    """Aplica estilo de encabezado a una fila."""
    fill = PatternFill(start_color='1E2333', end_color='1E2333', fill_type='solid')
    font = Font(bold=True, color='E8A020', size=10)
    for col, texto in enumerate(columnas, 1):
        cell = ws.cell(row=fila, column=col, value=texto)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')


def escribir_hoja(wb, titulo, encabezados, filas, ancho=18):
    """Agrega una hoja al workbook con encabezado estilado y filas ya formateadas.

    Reutiliza la hoja inicial vacía del Workbook en la primera llamada para
    no dejar una pestaña "Sheet" huérfana en workbooks multi-hoja.
    """
    if wb.active and wb.active.max_row == 1 and wb.active.max_column == 1 \
            and wb.active.cell(row=1, column=1).value is None:
        ws = wb.active
        ws.title = titulo
    else:
        ws = wb.create_sheet(title=titulo)

    estilo_encabezado(ws, 1, encabezados)
    for i, fila in enumerate(filas, 2):
        for col, valor in enumerate(fila, 1):
            ws.cell(row=i, column=col, value=valor)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = ancho
    return ws


def generar_reporte_excel(request, *, tipo, nombre_base, hojas,
                          fecha_inicio=None, fecha_fin=None, observaciones=''):
    """Genera el workbook, lo guarda en el historial y lo devuelve como descarga.

    hojas: lista de tuplas (titulo, encabezados, filas) o
           (titulo, encabezados, filas, ancho).

    El workbook se serializa UNA sola vez: los mismos bytes van al FileField
    (re-descarga desde el historial) y al cuerpo de la respuesta.
    """
    wb = openpyxl.Workbook()
    for hoja in hojas:
        titulo, encabezados, filas = hoja[0], hoja[1], hoja[2]
        ancho = hoja[3] if len(hoja) > 3 else 18
        escribir_hoja(wb, titulo, encabezados, filas, ancho)

    buffer = io.BytesIO()
    wb.save(buffer)
    datos = buffer.getvalue()

    nombre = f"{nombre_base}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    reporte = ReporteGenerado(
        tipo=tipo,
        generado_por=request.user,
        fecha_inicio_periodo=fecha_inicio or timezone.now().date(),
        fecha_fin_periodo=fecha_fin or timezone.now().date(),
        observaciones=observaciones,
    )
    reporte.archivo.save(nombre, ContentFile(datos), save=True)

    response = HttpResponse(
        datos,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre}"'
    return response


# ─────────────────────────────────────────────────────────────────────────────
# Consultas agregadas — fuente de datos del reporte RESUMEN (resumen ejecutivo)
# ─────────────────────────────────────────────────────────────────────────────

class ReporteService:

    @staticmethod
    def maquinas_mas_utilizadas(limite=10):
        from maquinas.models import Maquina
        return (Maquina.objects.exclude(estado='BAJA')
                .order_by('-horas_acumuladas')[:limite])

    @staticmethod
    def materiales_stock_bajo():
        from inventario.models import Material
        return (Material.objects.filter(activo=True,
                                        stock_actual__lte=F('stock_minimo'))
                .order_by('nombre'))

    @staticmethod
    def mantenimientos_pendientes():
        from mantenimiento.models import OrdenMantenimiento
        return (OrdenMantenimiento.objects
                .filter(activo=True, estado__in=['PROGRAMADA', 'EN_PROCESO'])
                .select_related('maquina', 'responsable_1')
                .order_by('fecha_programada'))

    @staticmethod
    def indicadores_tpm():
        from tpm.models import Alerta, CertificacionUsuario, Incidente, RegistroOEE

        hoy = timezone.now().date()

        oee_promedio = None
        oee_periodo = None
        ultimo = RegistroOEE.objects.filter(activo=True).order_by('-anio', '-mes').first()
        if ultimo:
            oee_periodo = f"{ultimo.mes:02d}/{ultimo.anio}"
            oee_promedio = (RegistroOEE.objects
                            .filter(activo=True, anio=ultimo.anio, mes=ultimo.mes)
                            .aggregate(prom=Avg('oee'))['prom'])

        alertas_activas = {
            severidad: Alerta.objects.filter(resuelta=False, severidad=severidad).count()
            for severidad, _ in Alerta.SEVERIDADES
        }

        certificaciones_por_vencer = (CertificacionUsuario.objects
                                      .filter(activo=True,
                                              fecha_vencimiento__gte=hoy,
                                              fecha_vencimiento__lte=hoy + timezone.timedelta(days=30))
                                      .select_related('usuario', 'maquina')
                                      .order_by('fecha_vencimiento'))

        incidentes_30d = Incidente.objects.filter(
            activo=True,
            fecha_ocurrencia__gte=timezone.now() - timezone.timedelta(days=30),
        ).count()

        return {
            'oee_promedio': oee_promedio,
            'oee_periodo': oee_periodo,
            'alertas_activas': alertas_activas,
            'certificaciones_por_vencer': certificaciones_por_vencer,
            'incidentes_30d': incidentes_30d,
        }


# ─────────────────────────────────────────────────────────────────────────────
# Constructores de hojas — cada uno devuelve (titulo, encabezados, filas[, ancho])
# y los reportes consolidados los combinan en un solo workbook.
# ─────────────────────────────────────────────────────────────────────────────

def hoja_materiales():
    from inventario.models import Material

    encabezados = ['Código', 'Nombre', 'Tipo', 'Unidad', 'Stock actual',
                   'Stock mínimo', 'Estado stock', 'Proveedor', 'Costo unitario']
    filas = [
        [m.codigo, m.nombre, m.get_tipo_display(), m.unidad_medida,
         float(m.stock_actual), float(m.stock_minimo),
         'Stock bajo' if m.stock_bajo else 'OK',
         fmt_texto(m.proveedor), float(m.costo_unitario)]
        for m in Material.objects.filter(activo=True)
    ]
    return ('Materiales', encabezados, filas, 18)


def hoja_consumos(fecha_inicio=None, fecha_fin=None):
    from inventario.models import ConsumoMaterial

    encabezados = ['Fecha', 'Material', 'Código', 'Unidad', 'Cantidad',
                   'OT', 'Registrado por', 'Observación']

    qs = ConsumoMaterial.objects.select_related('material', 'realizado_por', 'orden_trabajo')
    if fecha_inicio:
        qs = qs.filter(fecha__date__gte=fecha_inicio)
    if fecha_fin:
        qs = qs.filter(fecha__date__lte=fecha_fin)

    filas = [
        [fmt_fecha_hora(c.fecha), c.material.nombre, c.material.codigo,
         c.material.unidad_medida, float(c.cantidad),
         f'OT-{c.orden_trabajo.pk:04d}' if c.orden_trabajo else '—',
         fmt_usuario(c.realizado_por), fmt_texto(c.observacion)]
        for c in qs
    ]
    return ('Consumos', encabezados, filas, 18)


def hojas_piezas(fecha_inicio=None, fecha_fin=None):
    """Tres hojas: despiece actual, transferencias y reasignaciones.

    El despiece no se filtra por fechas (es el estado actual); las fechas
    aplican a transferencias y reasignaciones.
    """
    from maquinas.models import Pieza, TransferenciaPieza, ReasignacionPieza

    enc_piezas = ['Máquina', 'Ruta completa', 'Nombre', 'Nombre inglés',
                  'Nº parte', 'Posición', '¿Es ensamble?', 'Stock repuestos',
                  'Stock mínimo', 'Estado stock']
    filas_piezas = []
    piezas = Pieza.objects.filter(activo=True) \
                          .select_related('maquina', 'ensamble') \
                          .order_by('maquina__nombre', 'ensamble__nombre', 'nombre')
    for p in piezas:
        filas_piezas.append([
            p.maquina.nombre, p.get_ruta_completa(), p.nombre,
            fmt_texto(p.nombre_en), fmt_texto(p.numero_parte),
            p.numero_posicion if p.numero_posicion is not None else '—',
            si_no(p.es_ensamble), p.stock_repuestos, p.stock_minimo_repuestos,
            '—' if p.es_ensamble else ('Stock bajo' if p.stock_bajo else 'OK'),
        ])

    enc_transf = ['Fecha', 'Pieza', 'Nº parte', 'Máquina origen',
                  'Máquina destino', 'Autorizado por', 'Motivo', 'Observaciones']
    qs_transf = TransferenciaPieza.objects.select_related(
        'pieza', 'maquina_origen', 'maquina_destino', 'autorizado_por')
    if fecha_inicio:
        qs_transf = qs_transf.filter(fecha__date__gte=fecha_inicio)
    if fecha_fin:
        qs_transf = qs_transf.filter(fecha__date__lte=fecha_fin)
    filas_transf = [
        [fmt_fecha_hora(t.fecha), t.pieza.nombre, fmt_texto(t.pieza.numero_parte),
         t.maquina_origen.nombre if t.maquina_origen else '—',
         t.maquina_destino.nombre if t.maquina_destino else '—',
         fmt_usuario(t.autorizado_por), fmt_texto(t.motivo), fmt_texto(t.observaciones)]
        for t in qs_transf
    ]

    enc_reasig = ['Fecha', 'Pieza', 'Ensamble anterior', 'Ensamble nuevo', 'Realizado por']
    qs_reasig = ReasignacionPieza.objects.select_related(
        'pieza', 'ensamble_anterior', 'ensamble_nuevo', 'realizado_por')
    if fecha_inicio:
        qs_reasig = qs_reasig.filter(fecha__date__gte=fecha_inicio)
    if fecha_fin:
        qs_reasig = qs_reasig.filter(fecha__date__lte=fecha_fin)
    filas_reasig = [
        [fmt_fecha_hora(r.fecha), r.pieza.nombre,
         r.ensamble_anterior.nombre if r.ensamble_anterior else '(suelta)',
         r.ensamble_nuevo.nombre if r.ensamble_nuevo else '(suelta)',
         fmt_usuario(r.realizado_por)]
        for r in qs_reasig
    ]

    return [
        ('Piezas', enc_piezas, filas_piezas, 20),
        ('Transferencias', enc_transf, filas_transf, 20),
        ('Reasignaciones', enc_reasig, filas_reasig, 20),
    ]


def hoja_ordenes_mantenimiento(fecha_inicio=None, fecha_fin=None):
    from mantenimiento.models import OrdenMantenimiento

    encabezados = ['Nº', 'Máquina', 'Código', 'Tipo', 'Estado', 'Prioridad',
                   'Origen', 'Referencia origen', 'Título', 'Responsables',
                   'Fecha programada', 'Fecha inicio', 'Fecha fin',
                   'T. estimado (h)', 'Costo', '¿Afecta seguridad?',
                   '¿Para producción?', 'Autorizado por', 'Creado por']

    qs = OrdenMantenimiento.objects.select_related(
        'maquina', 'responsable_1', 'responsable_2', 'responsable_3',
        'autorizado_por', 'creado_por', 'plan', 'incidente', 'inspeccion',
        'hallazgo', 'parada', 'bitacora_operario',
    )
    if fecha_inicio:
        qs = qs.filter(fecha_programada__gte=fecha_inicio)
    if fecha_fin:
        qs = qs.filter(fecha_programada__lte=fecha_fin)

    def referencia_origen(om):
        # Los FKs de origen son SET_NULL: puede haber origen declarado sin referencia
        ref = (om.plan or om.incidente or om.inspeccion or om.hallazgo
               or om.parada or om.bitacora_operario)
        return str(ref) if ref else '—'

    filas = []
    for om in qs:
        responsables = '; '.join(
            fmt_usuario(r) for r in (om.responsable_1, om.responsable_2, om.responsable_3) if r
        ) or '—'
        filas.append([
            om.numero(), om.maquina.nombre, om.maquina.codigo,
            om.get_tipo_display(), om.get_estado_display(), om.get_prioridad_display(),
            om.get_origen_display(), referencia_origen(om), om.titulo, responsables,
            fmt_fecha(om.fecha_programada), fmt_fecha_hora(om.fecha_inicio),
            fmt_fecha_hora(om.fecha_fin), fmt_num(om.tiempo_estimado_horas),
            float(om.costo), si_no(om.afecta_seguridad), si_no(om.para_produccion),
            fmt_usuario(om.autorizado_por), fmt_usuario(om.creado_por),
        ])
    return ('Órdenes de mantenimiento', encabezados, filas, 20)


def hoja_mantenimientos(fecha_inicio=None, fecha_fin=None):
    from mantenimiento.models import Mantenimiento

    encabezados = ['#', 'Máquina', 'Código', 'Tipo', 'Estado', 'Prioridad',
                   'Fecha programada', 'Fecha inicio', 'Fecha fin',
                   'Horas trabajo', 'Costo', 'Responsable', 'Descripción']

    qs = Mantenimiento.objects.select_related('maquina', 'responsable')
    if fecha_inicio:
        qs = qs.filter(fecha_programada__gte=fecha_inicio)
    if fecha_fin:
        qs = qs.filter(fecha_programada__lte=fecha_fin)

    filas = [
        [m.pk, m.maquina.nombre, m.maquina.codigo, m.get_tipo_display(),
         m.get_estado_display(), m.get_prioridad_display(),
         fmt_fecha(m.fecha_programada), fmt_fecha_hora(m.fecha_inicio),
         fmt_fecha_hora(m.fecha_fin), float(m.horas_trabajo), float(m.costo),
         fmt_usuario(m.responsable), m.descripcion]
        for m in qs
    ]
    return ('Mantenimientos', encabezados, filas, 20)


def hoja_bitacora_mantenimiento(fecha_inicio=None, fecha_fin=None):
    from mantenimiento.models import BitacoraMantenimiento

    encabezados = ['Fecha', 'Máquina', 'OM', 'Técnico', 'Tipo actividad',
                   'Horas', 'Descripción', 'Repuestos utilizados', '¿Requiere atención?']

    qs = BitacoraMantenimiento.objects.select_related('maquina', 'tecnico', 'orden')
    if fecha_inicio:
        qs = qs.filter(fecha_registro__date__gte=fecha_inicio)
    if fecha_fin:
        qs = qs.filter(fecha_registro__date__lte=fecha_fin)

    filas = [
        [fmt_fecha_hora(b.fecha_registro), b.maquina.nombre,
         b.orden.numero() if b.orden else '—', fmt_usuario(b.tecnico),
         b.get_tipo_actividad_display() if b.tipo_actividad else '—',
         fmt_num(b.tiempo_horas), b.descripcion,
         fmt_texto(b.repuestos_utilizados), si_no(b.requiere_atencion)]
        for b in qs
    ]
    return ('Bitácora mantenimiento', encabezados, filas, 20)


def hoja_reservas(fecha_inicio=None, fecha_fin=None):
    from reservas.models import Reserva

    encabezados = ['#', 'Solicitante', 'Máquina', 'Código', 'Fecha',
                   'Hora inicio', 'Hora fin', 'Propósito', 'Estado',
                   'Autorizador', 'OT', 'T. planificado (min)', 'T. real (min)',
                   'T. paradas (min)', 'Unidades prod.', 'Unidades esp.']

    qs = Reserva.objects.select_related('usuario', 'maquina', 'autorizador', 'orden_trabajo')
    if fecha_inicio:
        qs = qs.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        qs = qs.filter(fecha__lte=fecha_fin)

    filas = []
    for r in qs:
        ot = getattr(r, 'orden_trabajo', None)
        filas.append([
            r.pk,
            fmt_usuario(r.usuario) if r.usuario else '— usuario eliminado —',
            r.maquina.nombre, r.maquina.codigo, fmt_fecha(r.fecha),
            r.hora_inicio.strftime('%H:%M'), r.hora_fin.strftime('%H:%M'),
            r.get_proposito_display(), r.get_estado_display(),
            fmt_usuario(r.autorizador),
            f'OT-{ot.pk:04d}' if ot else '—',
            float(ot.tiempo_planificado_min) if ot else '—',
            float(ot.tiempo_real_min) if ot and ot.tiempo_real_min else '—',
            float(ot.tiempo_parada_min) if ot else '—',
            ot.unidades_producidas if ot else '—',
            ot.unidades_esperadas if ot else '—',
        ])
    return ('Reservas y OT', encabezados, filas, 18)


def hoja_bitacora_operarios(fecha_inicio=None, fecha_fin=None):
    from reservas.models import BitacoraOperario

    encabezados = ['Fecha', 'OT', 'Máquina', 'Operario', 'Descripción',
                   'Observaciones', '¿Requiere atención?', '¿Tiene foto?']

    qs = BitacoraOperario.objects.select_related(
        'operario', 'orden_trabajo__reserva__maquina')
    if fecha_inicio:
        qs = qs.filter(fecha_registro__date__gte=fecha_inicio)
    if fecha_fin:
        qs = qs.filter(fecha_registro__date__lte=fecha_fin)

    filas = [
        [fmt_fecha_hora(b.fecha_registro), f'OT-{b.orden_trabajo.pk:04d}',
         b.orden_trabajo.reserva.maquina.nombre, fmt_usuario(b.operario),
         b.descripcion_trabajo, fmt_texto(b.observaciones),
         si_no(b.requiere_atencion), si_no(bool(b.foto))]
        for b in qs.order_by('-fecha_registro')
    ]
    return ('Bitácora operarios', encabezados, filas, 20)


def hoja_pareto(fecha_inicio=None, fecha_fin=None):
    from reservas.models import RegistroParada

    encabezados = ['Código parada', 'Tipo', 'Categoría', 'Subsistema',
                   'Frecuencia', 'Duración total (min)', '% Frecuencia acumulada']

    qs = RegistroParada.objects.filter(activo=True).select_related('codigo_parada')
    if fecha_inicio:
        qs = qs.filter(orden_trabajo__reserva__fecha__gte=fecha_inicio)
    if fecha_fin:
        qs = qs.filter(orden_trabajo__reserva__fecha__lte=fecha_fin)

    # Agrupar por código de parada
    agrupado = defaultdict(lambda: {'frecuencia': 0, 'duracion': 0, 'obj': None})
    for p in qs:
        key = p.codigo_parada.codigo if p.codigo_parada else 'SIN_CÓDIGO'
        agrupado[key]['frecuencia'] += 1
        agrupado[key]['duracion'] += float(p.duracion_minutos or 0)
        agrupado[key]['obj'] = p.codigo_parada

    # Ordenar por frecuencia descendente
    ordenado = sorted(agrupado.items(), key=lambda x: x[1]['frecuencia'], reverse=True)
    total = sum(v['frecuencia'] for _, v in ordenado)
    acumulado = 0

    filas = []
    for codigo, datos in ordenado:
        acumulado += datos['frecuencia']
        pct = round((acumulado / total * 100), 1) if total > 0 else 0
        obj = datos['obj']
        filas.append([
            codigo,
            obj.get_tipo_display() if obj else '—',
            obj.get_categoria_display() if obj else '—',
            obj.subsistema if obj else '—',
            datos['frecuencia'], round(datos['duracion'], 2), pct,
        ])
    return ('Pareto paradas', encabezados, filas, 22)


def hoja_oee(fecha_inicio=None, fecha_fin=None):
    """Los registros OEE son mensuales (anio/mes): el rango de fechas se
    traduce al mes calendario que contiene cada extremo."""
    from tpm.models import RegistroOEE

    encabezados = ['Período', 'Máquina', 'Código', 'Disponibilidad %',
                   'Rendimiento %', 'Calidad %', 'OEE %', 'Clasificación']

    registros = RegistroOEE.objects.select_related('maquina').all()
    if fecha_inicio:
        anio, mes = int(fecha_inicio[:4]), int(fecha_inicio[5:7])
        registros = registros.filter(Q(anio__gt=anio) | Q(anio=anio, mes__gte=mes))
    if fecha_fin:
        anio, mes = int(fecha_fin[:4]), int(fecha_fin[5:7])
        registros = registros.filter(Q(anio__lt=anio) | Q(anio=anio, mes__lte=mes))
    registros = registros.order_by('anio', 'mes', 'maquina__nombre')

    filas = []
    for r in registros:
        oee = float(r.oee)
        filas.append([
            f"{r.mes:02d}/{r.anio}", r.maquina.nombre, r.maquina.codigo,
            float(r.disponibilidad), float(r.rendimiento), float(r.calidad), oee,
            'World class' if oee >= 85 else 'Aceptable' if oee >= 60 else 'Mejorar',
        ])
    return ('OEE', encabezados, filas, 18)


def hoja_inspecciones(fecha_inicio=None, fecha_fin=None):
    from tpm.models import InspeccionDiaria

    encabezados = ['Fecha', 'Máquina', 'Inspector', 'Limpieza', 'Temperatura', 'Guardas',
                   'Emergencia', 'Ruidos', 'Vibraciones', 'Ítems específicos (catálogo)',
                   'Resultado', 'Observaciones']

    qs = InspeccionDiaria.objects.select_related('maquina', 'inspector') \
                                 .prefetch_related('respuestas_checklist__item')
    if fecha_inicio:
        qs = qs.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        qs = qs.filter(fecha__lte=fecha_fin)

    ok_falla = lambda v: 'OK' if v else 'FALLA'

    filas = []
    for insp in qs:
        items_especificos = '; '.join(
            f"{r.item.nombre}: {'OK' if r.ok else 'FALLA'}" for r in insp.respuestas_checklist.all()
        ) or '—'
        filas.append([
            fmt_fecha(insp.fecha), insp.maquina.nombre, fmt_usuario(insp.inspector),
            ok_falla(insp.limpieza_area_ok), ok_falla(insp.temperatura_normal),
            ok_falla(insp.guardas_seguridad_ok), ok_falla(insp.boton_emergencia_ok),
            si_no(insp.ruidos_anormales), si_no(insp.vibraciones_anormales),
            items_especificos, 'Aprobada' if insp.aprobada else 'FALLIDA',
            fmt_texto(insp.observaciones),
        ])
    return ('Inspecciones', encabezados, filas, 16)


def hoja_incidentes(fecha_inicio=None, fecha_fin=None):
    from tpm.models import Incidente

    encabezados = ['Fecha ocurrencia', 'Máquina', 'Código', 'Reportado por',
                   'Tipo', 'Severidad', 'Descripción', 'Acción tomada',
                   '¿Requiere mantenimiento?', 'OMs generadas', '¿Activo?']

    qs = Incidente.objects.select_related('maquina', 'reportado_por') \
                          .prefetch_related('ordenes_generadas')
    if fecha_inicio:
        qs = qs.filter(fecha_ocurrencia__date__gte=fecha_inicio)
    if fecha_fin:
        qs = qs.filter(fecha_ocurrencia__date__lte=fecha_fin)

    filas = []
    for inc in qs:
        oms = '; '.join(om.numero() for om in inc.ordenes_generadas.all()) or '—'
        filas.append([
            fmt_fecha_hora(inc.fecha_ocurrencia), inc.maquina.nombre,
            inc.maquina.codigo, fmt_usuario(inc.reportado_por),
            inc.get_tipo_display(), inc.get_severidad_display(),
            inc.descripcion, fmt_texto(inc.accion_tomada),
            si_no(inc.requiere_mantenimiento), oms, si_no(inc.activo),
        ])
    return ('Incidentes', encabezados, filas, 20)


def hoja_alertas(fecha_inicio=None, fecha_fin=None):
    from tpm.models import Alerta

    encabezados = ['ID', 'Tipo', 'Severidad', 'Máquina', 'Mensaje', 'Generada',
                   'Asignada a', 'Asignada en', 'Vista por', 'Vista en',
                   '¿Resuelta?', 'Resuelta por', 'Resuelta en',
                   'T. resolución (h)', 'Nota resolución']

    qs = Alerta.objects.select_related('maquina', 'asignado_a', 'vista_por', 'resuelta_por')
    if fecha_inicio:
        qs = qs.filter(generada_en__date__gte=fecha_inicio)
    if fecha_fin:
        qs = qs.filter(generada_en__date__lte=fecha_fin)

    filas = []
    for a in qs:
        horas_resolucion = (
            round((a.resuelta_en - a.generada_en).total_seconds() / 3600, 1)
            if a.resuelta_en else '—'
        )
        filas.append([
            a.pk, a.get_tipo_display(), a.get_severidad_display(),
            a.maquina.nombre if a.maquina else '—', a.mensaje,
            fmt_fecha_hora(a.generada_en),
            fmt_usuario(a.asignado_a), fmt_fecha_hora(a.asignado_en),
            fmt_usuario(a.vista_por), fmt_fecha_hora(a.vista_en),
            si_no(a.resuelta), fmt_usuario(a.resuelta_por),
            fmt_fecha_hora(a.resuelta_en), horas_resolucion,
            fmt_texto(a.nota_resolucion),
        ])
    return ('Alertas', encabezados, filas, 18)


def hoja_certificaciones():
    from tpm.models import CertificacionUsuario

    encabezados = ['Usuario', 'Máquina', 'Código', 'Otorgada por',
                   'Fecha otorgamiento', 'Fecha vencimiento', '¿Vigente?',
                   'Días para vencer', '¿Activa?', 'Observaciones']

    filas = [
        [fmt_usuario(c.usuario), c.maquina.nombre, c.maquina.codigo,
         fmt_usuario(c.otorgado_por), fmt_fecha(c.fecha_otorgamiento),
         fmt_fecha(c.fecha_vencimiento), si_no(c.vigente), c.dias_para_vencer,
         si_no(c.activo), fmt_texto(c.observaciones)]
        for c in CertificacionUsuario.objects.select_related('usuario', 'maquina', 'otorgado_por')
    ]
    return ('Certificaciones', encabezados, filas, 18)


def hojas_resumen():
    """Seis hojas del resumen ejecutivo (KPIs y focos de atención)."""
    from maquinas.models import Maquina, Pieza
    from mantenimiento.models import PlanMantenimiento
    from tpm.models import Alerta

    hoy = timezone.now().date()

    pendientes = ReporteService.mantenimientos_pendientes()
    stock_bajo = ReporteService.materiales_stock_bajo()
    mas_usadas = ReporteService.maquinas_mas_utilizadas()
    tpm = ReporteService.indicadores_tpm()

    # Hoja 1 — KPIs generales
    enc_kpis = ['Indicador', 'Valor']
    filas_kpis = [['Total de máquinas', Maquina.objects.count()]]
    for estado, etiqueta in Maquina.ESTADOS:
        filas_kpis.append([f'Máquinas — {etiqueta}',
                           Maquina.objects.filter(estado=estado).count()])
    oms_vencidas = sum(1 for om in pendientes if om.fecha_programada < hoy)
    filas_kpis += [
        ['Órdenes de mantenimiento abiertas', len(pendientes)],
        ['Órdenes de mantenimiento vencidas', oms_vencidas],
        ['Planes de mantenimiento activos',
         PlanMantenimiento.objects.filter(activo=True).count()],
        ['Materiales en stock bajo', stock_bajo.count()],
        ['Piezas de repuesto en stock bajo',
         Pieza.objects.filter(activo=True, es_ensamble=False,
                              stock_repuestos__lte=F('stock_minimo_repuestos')).count()],
        [f"OEE promedio ({tpm['oee_periodo'] or 'sin registros'})",
         round(float(tpm['oee_promedio']), 2) if tpm['oee_promedio'] is not None else '—'],
    ]
    for severidad, etiqueta in Alerta.SEVERIDADES:
        filas_kpis.append([f'Alertas sin resolver — {etiqueta}',
                           tpm['alertas_activas'].get(severidad, 0)])
    filas_kpis += [
        ['Certificaciones por vencer (≤30 días)', tpm['certificaciones_por_vencer'].count()],
        ['Incidentes en los últimos 30 días', tpm['incidentes_30d']],
    ]

    # Hoja 2 — Máquinas más utilizadas
    enc_maquinas = ['Máquina', 'Código', 'Estado', 'Horas acumuladas']
    filas_maquinas = [
        [m.nombre, m.codigo, m.get_estado_display(), float(m.horas_acumuladas)]
        for m in mas_usadas
    ]

    # Hoja 3 — Materiales en stock bajo
    enc_stock = ['Código', 'Material', 'Stock actual', 'Stock mínimo', 'Unidad']
    filas_stock = [
        [m.codigo, m.nombre, float(m.stock_actual), float(m.stock_minimo), m.unidad_medida]
        for m in stock_bajo
    ]

    # Hoja 4 — Mantenimientos pendientes (OMs abiertas)
    enc_pend = ['Nº', 'Máquina', 'Título', 'Prioridad', 'Fecha programada',
                '¿Vencida?', 'Responsable']
    filas_pend = [
        [om.numero(), om.maquina.nombre, om.titulo, om.get_prioridad_display(),
         fmt_fecha(om.fecha_programada), si_no(om.fecha_programada < hoy),
         fmt_usuario(om.responsable_1)]
        for om in pendientes
    ]

    # Hoja 5 — Alertas activas
    enc_alertas = ['Tipo', 'Severidad', 'Máquina', 'Mensaje', 'Generada']
    filas_alertas = [
        [a.get_tipo_display(), a.get_severidad_display(),
         a.maquina.nombre if a.maquina else '—', a.mensaje,
         fmt_fecha_hora(a.generada_en)]
        for a in Alerta.objects.filter(resuelta=False).select_related('maquina')
    ]

    # Hoja 6 — Certificaciones por vencer
    enc_cert = ['Usuario', 'Máquina', 'Vence', 'Días restantes']
    filas_cert = [
        [fmt_usuario(c.usuario), c.maquina.nombre,
         fmt_fecha(c.fecha_vencimiento), c.dias_para_vencer]
        for c in tpm['certificaciones_por_vencer']
    ]

    return [
        ('KPIs', enc_kpis, filas_kpis, 34),
        ('Máquinas más utilizadas', enc_maquinas, filas_maquinas, 18),
        ('Stock bajo', enc_stock, filas_stock, 18),
        ('Mantenimientos pendientes', enc_pend, filas_pend, 20),
        ('Alertas activas', enc_alertas, filas_alertas, 22),
        ('Certificaciones por vencer', enc_cert, filas_cert, 20),
    ]


# ─────────────────────────────────────────────────────────────────────────────
# Respaldo completo — volcado genérico de todas las tablas del sistema.
# Una hoja por modelo (más las tablas intermedias de los M2M), con los valores
# crudos de cada columna (FKs como *_id), para servir de copia de seguridad.
# ─────────────────────────────────────────────────────────────────────────────

APPS_RESPALDO = ['usuarios', 'maquinas', 'inventario', 'mantenimiento',
                 'reservas', 'tpm', 'reportes']

# Nunca volcar hashes de contraseñas a un Excel descargable
CAMPOS_EXCLUIDOS_RESPALDO = {'password'}


def _valor_respaldo(valor):
    """Convierte un valor crudo de la BD a algo que openpyxl acepte."""
    if valor is None:
        return None
    if isinstance(valor, dt.datetime):
        if tz.is_aware(valor):
            valor = tz.localtime(valor)
        return valor.replace(tzinfo=None)
    if isinstance(valor, (bool, int, float, Decimal, dt.date, dt.time)):
        return valor
    # FieldFile, timedelta, UUID, JSON, etc. → texto sin caracteres ilegales
    return ILLEGAL_CHARACTERS_RE.sub('', str(valor))


def _titulo_hoja_unico(nombre, usados):
    """Excel limita los títulos de hoja a 31 caracteres y exige unicidad."""
    titulo = nombre[:31]
    base, n = titulo, 2
    while titulo in usados:
        sufijo = f'_{n}'
        titulo = base[:31 - len(sufijo)] + sufijo
        n += 1
    usados.add(titulo)
    return titulo


def hojas_backup():
    from django.apps import apps as registro_apps

    hojas = []
    indice = []
    titulos_usados = {'Índice'}
    modelos_vistos = set()

    def volcar_modelo(modelo, app_label):
        campos = [f for f in modelo._meta.concrete_fields
                  if f.name not in CAMPOS_EXCLUIDOS_RESPALDO]
        encabezados = [f.column for f in campos]
        # _base_manager: ignora managers que filtren (queremos TODO, incl. inactivos)
        filas = [
            [_valor_respaldo(getattr(obj, f.attname)) for f in campos]
            for obj in modelo._base_manager.all().order_by('pk').iterator()
        ]
        titulo = _titulo_hoja_unico(modelo.__name__, titulos_usados)
        indice.append([app_label, titulo, modelo._meta.db_table, len(filas)])
        hojas.append((titulo, encabezados, filas, 16))

    for app_label in APPS_RESPALDO:
        for modelo in registro_apps.get_app_config(app_label).get_models():
            if modelo in modelos_vistos:
                continue
            modelos_vistos.add(modelo)
            volcar_modelo(modelo, app_label)
            # Tablas intermedias de los ManyToMany (p. ej. roles/permisos)
            for campo_m2m in modelo._meta.local_many_to_many:
                through = campo_m2m.remote_field.through
                if through in modelos_vistos:
                    continue
                modelos_vistos.add(through)
                volcar_modelo(through, app_label)

    return [('Índice', ['App', 'Hoja', 'Tabla en BD', 'Registros'], indice, 24)] + hojas
